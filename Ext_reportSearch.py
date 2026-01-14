import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import openpyxl
from openpyxl import load_workbook
import time
import threading

class ExcelAccountFinder:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Account Finder")
        self.root.geometry("800x600")
        self.root.resizable(True, True)
        
        # Variables
        self.file_paths = []
        self.results = []
        
        self.setup_ui()
    
    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(4, weight=1)
        
        # Title
        ttk.Label(main_frame, text="Excel Account Finder", 
                 font=('Arial', 16, 'bold')).grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # File selection
        ttk.Label(main_frame, text="Excel Files:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.file_label = ttk.Label(main_frame, text="No files selected", 
                                   foreground="gray")
        self.file_label.grid(row=1, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Button(main_frame, text="Select Excel Files", 
                  command=self.select_files).grid(row=2, column=0, pady=10)
        
        # Sheets display
        ttk.Label(main_frame, text="Available Sheets:").grid(row=3, column=0, sticky=tk.W, pady=(20,5))
        self.sheets_listbox = tk.Listbox(main_frame, height=4)
        self.sheets_listbox.grid(row=3, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=(20,5))
        
        ttk.Button(main_frame, text="Select All Sheets", 
                  command=self.select_all_sheets).grid(row=4, column=0, pady=5)
        ttk.Button(main_frame, text="Select Sheets", 
                  command=self.select_sheets).grid(row=4, column=1, pady=5)
        
        # Account number input
        ttk.Label(main_frame, text="Account Number:").grid(row=5, column=0, sticky=tk.W, pady=(20,5))
        self.account_entry = ttk.Entry(main_frame, font=('Arial', 12), width=20)
        self.account_entry.grid(row=5, column=1, sticky=tk.W, pady=(20,5))
        self.account_entry.bind('<Return>', lambda e: self.search_account())
        
        # Search button
        ttk.Button(main_frame, text="🔍 Search Account", 
                  command=self.search_account).grid(row=5, column=2, padx=(10,0), pady=(20,5))
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        # Results display
        ttk.Label(main_frame, text="Results:").grid(row=7, column=0, sticky=(tk.W, tk.N), pady=(20,5))
        self.results_text = scrolledtext.ScrolledText(main_frame, height=15, width=80)
        self.results_text.grid(row=8, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # Save buttons frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=9, column=0, columnspan=3, pady=10)
        btn_frame.columnconfigure(1, weight=1)
        
        ttk.Button(btn_frame, text="💾 Save to Master File", 
                  command=self.save_results).grid(row=0, column=0, padx=5)
        ttk.Button(btn_frame, text="📋 Copy Results", 
                  command=self.copy_results).grid(row=0, column=1, padx=5)
        ttk.Button(btn_frame, text="🗑 Clear", 
                  command=self.clear_results).grid(row=0, column=2, padx=5)
    
    def select_files(self):
        filetypes = [("Excel files", "*.xlsx *.xls")]
        selected = filedialog.askopenfilenames(title="Select Excel Files", filetypes=filetypes)
        if selected:
            self.file_paths = list(selected)
            file_names = [os.path.basename(f) for f in self.file_paths]
            self.file_label.config(text=f"{len(self.file_paths)} files selected", foreground="green")
            
            # Get all unique sheets
            all_sheets = set()
            for file_path in self.file_paths:
                try:
                    xls = pd.ExcelFile(file_path)
                    all_sheets.update(xls.sheet_names)
                except Exception as e:
                    messagebox.showerror("Error", f"Error reading {os.path.basename(file_path)}: {e}")
            
            # Update sheets listbox
            self.sheets_listbox.delete(0, tk.END)
            for sheet in sorted(all_sheets):
                self.sheets_listbox.insert(tk.END, sheet)
    
    def select_all_sheets(self):
        sheets = list(self.sheets_listbox.get(0, tk.END))
        self.sheets_listbox.selection_set(0, tk.END)
    
    def select_sheets(self):
        # This would open a dialog to select specific sheets
        sheets = list(self.sheets_listbox.get(0, tk.END))
        if sheets:
            self.sheets_listbox.selection_set(0, tk.END)
    
    def search_account(self):
        account_number = self.account_entry.get().strip()
        if not account_number:
            messagebox.showwarning("Warning", "Please enter an account number")
            return
        
        if not self.file_paths:
            messagebox.showwarning("Warning", "Please select Excel files first")
            return
        
        # Run search in separate thread
        self.progress.start()
        thread = threading.Thread(target=self._perform_search, args=(account_number,))
        thread.daemon = True
        thread.start()
    
    def _perform_search(self, account_number):
        try:
            sheets_to_search = list(self.sheets_listbox.get(0, tk.END))
            if not sheets_to_search:
                sheets_to_search = ["Sheet1"]  # Default
            
            self.results = []
            
            for file_path in self.file_paths:
                self.root.after(0, lambda fp=file_path: self.update_status(f"Searching: {os.path.basename(fp)}"))
                for sheet in sheets_to_search:
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet)
                        if 'Account No' in df.columns:
                            matched_rows = df[df['Account No'] == account_number]
                            for _, row in matched_rows.iterrows():
                                self.results.append({
                                    "File": os.path.basename(file_path),
                                    "Sheet": sheet,
                                    "Letter Ref": row.get("Letter Ref", ""),
                                    "Letter Date": row.get("Letter Date", "")
                                })
                    except ValueError:
                        pass
                    except Exception as e:
                        pass
            
            # Update UI on main thread
            self.root.after(0, self._display_results)
            #exception to handle the unwanted stuff
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", str(e)))
        finally:
            self.root.after(0, self.progress.stop)
    
    def _display_results(self):
        self.results_text.delete(1.0, tk.END)
        if self.results:
            result_df = pd.DataFrame(self.results)
            display_text = result_df.to_string(index=False)
            self.results_text.insert(tk.END, f"✅ Account Found ({len(self.results)} matches):\n\n{display_text}")
        else:
            self.results_text.insert(tk.END, "❌ Account number not found in the selected sheets.")
    
    def update_status(self, message):
        self.results_text.insert(tk.END, f"\n{message}")
        self.results_text.see(tk.END)
        self.root.update()
    
    def save_results(self):
        if not self.results:
            messagebox.showwarning("Warning", "No results to save")
            return
        
        result_df = pd.DataFrame(self.results)
        master_file = "master_account_results.xlsx"
        
        max_retries = 5
        for attempt in range(max_retries):
            try:
                book = load_workbook(master_file, data_only=False)
                if book.sheetnames:
                    sheet = book[book.sheetnames[0]]
                else:
                    sheet = book.create_sheet("Results")
                
                new_rows = result_df.values.tolist()
                next_row = sheet.max_row + 1 if sheet.max_row > 0 else 1
                
                for row_data in new_rows:
                    sheet.append(row_data)
                
                book.save(master_file)
                messagebox.showinfo("Success", f"✅ Data added to '{master_file}' (Rows {next_row}-{next_row+len(new_rows)-1})")
                return
                
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(2)
                else:
                    temp_file = f"master_account_results_temp_{int(time.time())}.xlsx"
                    result_df.to_excel(temp_file, index=False)
                    messagebox.showinfo("Saved", f"💾 Saved to '{temp_file}' (Excel was locked)")
                    return
            except FileNotFoundError:
                result_df.to_excel(master_file, index=False)
                messagebox.showinfo("Success", f"✅ New file created: '{master_file}'")
                return
            except Exception as e:
                if attempt < max_retries - 1:
                    time.sleep(2)
                else:
                    temp_file = f"master_account_results_temp_{int(time.time())}.xlsx"
                    result_df.to_excel(temp_file, index=False)
                    messagebox.showinfo("Saved", f"💾 Fallback saved to '{temp_file}'")
                    return
    
    def copy_results(self):
        if self.results:
            result_df = pd.DataFrame(self.results)
            self.root.clipboard_clear()
            self.root.clipboard_append(result_df.to_csv(sep='\t', index=False))
            messagebox.showinfo("Copied", "Results copied to clipboard")
    
    def clear_results(self):
        self.results_text.delete(1.0, tk.END)
        self.account_entry.delete(0, tk.END)
        self.results = []

def main():
    root = tk.Tk()
    app = ExcelAccountFinder(root)
    root.mainloop()

if __name__ == "__main__":
    main()
